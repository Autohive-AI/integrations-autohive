from autohive_integrations_sdk import (
    Integration, ExecutionContext, ActionHandler
)
from typing import Dict, Any, List, Optional
import base64
import io
import json
import re
import csv

spreadsheet_tools = Integration.load()


def sanitize_header(header: str, existing_headers: set) -> str:
    """
    Sanitize a header string to be a valid JSON property name.
    - Remove/replace invalid characters
    - Ensure it doesn't start with a number
    - Handle duplicates by adding suffixes
    """
    if not header or not str(header).strip():
        header = "column"
    
    header = str(header).strip()
    
    # Replace spaces and special characters with underscores
    sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', header)
    
    # Remove consecutive underscores
    sanitized = re.sub(r'_+', '_', sanitized)
    
    # Remove leading/trailing underscores
    sanitized = sanitized.strip('_')
    
    # Ensure it doesn't start with a number
    if sanitized and sanitized[0].isdigit():
        sanitized = f"col_{sanitized}"
    
    # If empty after sanitization, use default
    if not sanitized:
        sanitized = "column"
    
    # Handle duplicates
    original_sanitized = sanitized
    counter = 1
    while sanitized in existing_headers:
        sanitized = f"{original_sanitized}_{counter}"
        counter += 1
    
    return sanitized


def read_csv_data(file_bytes: bytes) -> tuple[List[str], List[List[Any]]]:
    """Read CSV file and return headers and rows"""
    text_content = file_bytes.decode('utf-8-sig')  # utf-8-sig handles BOM
    reader = csv.reader(io.StringIO(text_content))
    
    rows = list(reader)
    if not rows:
        raise ValueError("CSV file is empty")
    
    headers = rows[0]
    data_rows = rows[1:]
    
    return headers, data_rows


def read_excel_xlsx(file_bytes: bytes) -> tuple[List[str], List[List[Any]]]:
    """Read Excel .xlsx file and return headers and rows"""
    from openpyxl import load_workbook
    
    workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")
    
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    data_rows = [[cell for cell in row] for row in rows[1:]]
    
    workbook.close()
    return headers, data_rows


def read_excel_xls(file_bytes: bytes) -> tuple[List[str], List[List[Any]]]:
    """Read Excel .xls file and return headers and rows"""
    import xlrd
    
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)
    
    if sheet.nrows == 0:
        raise ValueError("Excel file is empty")
    
    headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
    data_rows = []
    
    for row_idx in range(1, sheet.nrows):
        row = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
        data_rows.append(row)
    
    return headers, data_rows


def convert_to_typed_value(value: Any) -> Any:
    """Convert value to appropriate JSON type"""
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, str):
        # Try to convert numeric strings
        try:
            if '.' in value:
                return float(value)
            else:
                return int(value)
        except (ValueError, AttributeError):
            return value
    return str(value)


@spreadsheet_tools.action("convert_to_json")
class ConvertToJsonAction(ActionHandler):
    async def execute(self, inputs: Dict[str, Any], context: ExecutionContext):
        file_info = inputs.get("file")
        if not file_info:
            raise ValueError("File is required")
        
        file_name = file_info.get("name", "")
        file_content_b64 = file_info.get("content", "")
        content_type = file_info.get("contentType", "")
        
        if not file_content_b64:
            raise ValueError("File content is required")
        
        # Decode the file
        try:
            file_bytes = base64.b64decode(file_content_b64)
        except Exception as e:
            raise ValueError(f"Failed to decode file content: {str(e)}")
        
        # Determine file type and read data
        file_extension = file_name.lower().split('.')[-1] if '.' in file_name else ''
        
        headers = None
        data_rows = None
        parse_error = None
        
        if file_extension == 'csv' or content_type == 'text/csv':
            try:
                headers, data_rows = read_csv_data(file_bytes)
            except Exception as e:
                parse_error = f"Failed to parse CSV file: {str(e)}"
        elif file_extension == 'xlsx' or 'openxmlformats' in content_type:
            try:
                headers, data_rows = read_excel_xlsx(file_bytes)
            except Exception as e:
                parse_error = f"Failed to parse Excel (.xlsx) file: {str(e)}"
        elif file_extension == 'xls':
            try:
                headers, data_rows = read_excel_xls(file_bytes)
            except Exception as e:
                parse_error = f"Failed to parse Excel (.xls) file: {str(e)}"
        else:
            # Try to guess - attempt CSV first, then Excel
            try:
                headers, data_rows = read_csv_data(file_bytes)
            except:
                try:
                    headers, data_rows = read_excel_xlsx(file_bytes)
                except:
                    try:
                        headers, data_rows = read_excel_xls(file_bytes)
                    except Exception as e:
                        parse_error = f"Unable to parse file as spreadsheet. Supported formats: Excel (.xlsx, .xls), CSV (.csv). Error: {str(e)}"
        
        if headers is None or data_rows is None:
            if parse_error:
                raise ValueError(parse_error)
            else:
                raise ValueError("Failed to parse file as a spreadsheet. The file may not contain valid spreadsheet data.")
        
        # Sanitize headers
        sanitized_headers = []
        seen_headers = set()
        for header in headers:
            sanitized = sanitize_header(header, seen_headers)
            sanitized_headers.append(sanitized)
            seen_headers.add(sanitized)
        
        # Convert rows to JSON objects
        json_data = []
        for row in data_rows:
            row_obj = {}
            for idx, header in enumerate(sanitized_headers):
                value = row[idx] if idx < len(row) else None
                row_obj[header] = convert_to_typed_value(value)
            json_data.append(row_obj)
        
        # Convert to JSON string with proper formatting
        json_string = json.dumps(json_data, indent=2, ensure_ascii=False)
        
        # Encode to base64
        json_bytes = json_string.encode('utf-8')
        content_b64 = base64.b64encode(json_bytes).decode('utf-8')
        
        # Generate output filename
        base_name = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name
        output_filename = f"{base_name}.json"
        
        return {
            "file": {
                "content": content_b64,
                "name": output_filename,
                "contentType": "application/json"
            }
        }
